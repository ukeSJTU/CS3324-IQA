"""
Dataset folder classes for various IQA databases.

This module provides PyTorch Dataset implementations for loading
images and quality labels from different IQA benchmark databases.
"""

from __future__ import annotations

import csv
import os
from abc import ABC, abstractmethod
from typing import Any, Callable

import numpy as np
import scipy.io
import torch.utils.data as data
from openpyxl import load_workbook
from PIL import Image

# LIVE database distortion type image counts
LIVE_JP2K_COUNT = 227
LIVE_JPEG_COUNT = 233
LIVE_WN_COUNT = 174
LIVE_GBLUR_COUNT = 174
LIVE_FASTFADING_COUNT = 174

# LIVE Challenge database slice indices (excluding training images)
LIVEC_START_INDEX = 7
LIVEC_END_INDEX = 1169

# BID database total image count
BID_IMAGE_COUNT = 586


def pil_loader(path: str) -> Image.Image:
    """
    Load an image from disk using PIL.

    Args:
        path: Path to the image file.

    Returns:
        PIL Image in RGB format.
    """
    with open(path, "rb") as f:
        img = Image.open(f)
        return img.convert("RGB")


def get_filenames_by_suffix(directory: str, suffix: str) -> list[str]:
    """
    Get all filenames in a directory with the specified suffix.

    Args:
        directory: Directory path to search.
        suffix: File extension to filter by (e.g., '.bmp').

    Returns:
        List of matching filenames.
    """
    filenames = []
    for filename in os.listdir(directory):
        if os.path.splitext(filename)[1] == suffix:
            filenames.append(filename)
    return filenames


def get_tid_reference_names(directory: str, suffixes: str) -> list[str]:
    """
    Get reference image identifiers from TID2013 database.

    Args:
        directory: Path to reference images directory.
        suffixes: String containing valid suffixes (e.g., '.bmp.BMP').

    Returns:
        List of two-character reference image identifiers.
    """
    identifiers = []
    for filename in os.listdir(directory):
        ext = os.path.splitext(filename)[1]
        if ext in suffixes:
            identifiers.append(filename[1:3])
    return identifiers


class BaseIQADataset(data.Dataset, ABC):
    """
    Abstract base class for IQA dataset folders.

    Provides common functionality for loading and transforming images
    with their quality labels.
    """

    def __init__(
        self,
        root: str,
        index: list[int],
        transform: Callable[[Image.Image], Any] | None,
        patch_num: int,
    ) -> None:
        """
        Initialize the dataset.

        Args:
            root: Root directory of the dataset.
            index: List of indices to include in this dataset split.
            transform: Optional transform to apply to images.
            patch_num: Number of patches/augmentations per image.
        """
        self.root = root
        self.transform = transform
        self.patch_num = patch_num
        self.samples: list[tuple[str, np.floating[Any]]] = []
        self._build_samples(index)

    @abstractmethod
    def _build_samples(self, index: list[int]) -> None:
        """Build the list of (image_path, label) samples."""
        pass

    def __getitem__(self, index: int) -> tuple[Any, np.floating[Any]]:
        """
        Get a sample by index.

        Args:
            index: Sample index.

        Returns:
            Tuple of (transformed_image, quality_label).
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        if self.transform is not None:
            sample = self.transform(sample)
        return sample, target

    def __len__(self) -> int:
        """Return the total number of samples."""
        return len(self.samples)


class LIVEFolder(BaseIQADataset):
    """
    Dataset class for the LIVE IQA database.

    The LIVE database contains images with five distortion types:
    JP2K compression, JPEG compression, white noise, Gaussian blur,
    and fast fading (simulated Rayleigh fading channel).
    """

    def _get_distortion_filenames(self, path: str, count: int) -> list[str]:
        """
        Generate numbered image filenames for a distortion type.

        Args:
            path: Directory containing distorted images.
            count: Number of images in this distortion category.

        Returns:
            List of full paths to distorted images.
        """
        return [os.path.join(path, f"img{i + 1}.bmp") for i in range(count)]

    def _build_samples(self, index: list[int]) -> None:
        """Build samples from LIVE database structure."""
        # Load reference image names
        ref_path = os.path.join(self.root, "refimgs")
        ref_names = get_filenames_by_suffix(ref_path, ".bmp")

        # Collect all distorted image paths
        distortion_configs = [
            ("jp2k", LIVE_JP2K_COUNT),
            ("jpeg", LIVE_JPEG_COUNT),
            ("wn", LIVE_WN_COUNT),
            ("gblur", LIVE_GBLUR_COUNT),
            ("fastfading", LIVE_FASTFADING_COUNT),
        ]

        all_image_paths: list[str] = []
        for dist_type, count in distortion_configs:
            dist_path = os.path.join(self.root, dist_type)
            all_image_paths.extend(self._get_distortion_filenames(dist_path, count))

        # Load labels and reference mappings
        dmos_data = scipy.io.loadmat(os.path.join(self.root, "dmos_realigned.mat"))
        labels = dmos_data["dmos_new"].astype(np.float32)
        orgs = dmos_data["orgs"]

        refnames_data = scipy.io.loadmat(os.path.join(self.root, "refnames_all.mat"))
        refnames_all = refnames_data["refnames_all"]

        # Build samples based on train/test split
        for idx in index:
            # Find images corresponding to this reference
            matches = ref_names[idx] == refnames_all
            # Exclude original (non-distorted) images
            matches = matches & ~orgs.astype(bool)
            matching_indices = np.where(matches)[1].tolist()

            for img_idx in matching_indices:
                for _ in range(self.patch_num):
                    self.samples.append((all_image_paths[img_idx], labels[0][img_idx]))


class LIVEChallengeFolder(BaseIQADataset):
    """
    Dataset class for the LIVE Challenge (LIVEC) database.

    Contains authentically distorted images captured with mobile devices.
    """

    def _build_samples(self, index: list[int]) -> None:
        """Build samples from LIVE Challenge database."""
        # Load image paths and labels
        imgpath_data = scipy.io.loadmat(
            os.path.join(self.root, "Data", "AllImages_release.mat")
        )
        all_imgpaths = imgpath_data["AllImages_release"]
        all_imgpaths = all_imgpaths[LIVEC_START_INDEX:LIVEC_END_INDEX]

        mos_data = scipy.io.loadmat(
            os.path.join(self.root, "Data", "AllMOS_release.mat")
        )
        all_labels = mos_data["AllMOS_release"].astype(np.float32)
        all_labels = all_labels[0][LIVEC_START_INDEX:LIVEC_END_INDEX]

        # Build samples
        for idx in index:
            img_name = all_imgpaths[idx][0][0]
            img_path = os.path.join(self.root, "Images", img_name)
            for _ in range(self.patch_num):
                self.samples.append((img_path, all_labels[idx]))


class CSIQFolder(BaseIQADataset):
    """
    Dataset class for the CSIQ IQA database.

    Contains images with six distortion types at multiple quality levels.
    """

    def _build_samples(self, index: list[int]) -> None:
        """Build samples from CSIQ database."""
        # Load reference image names
        ref_path = os.path.join(self.root, "src_imgs")
        ref_names = get_filenames_by_suffix(ref_path, ".png")

        # Parse label file
        txt_path = os.path.join(self.root, "csiq_label.txt")
        img_names: list[str] = []
        labels_list: list[str] = []
        refnames_all: list[str] = []

        with open(txt_path, "r") as f:
            for line in f:
                parts = line.strip().split()
                if len(parts) >= 2:
                    img_names.append(parts[0])
                    labels_list.append(parts[1])
                    # Extract reference name from distorted image name
                    name_parts = parts[0].split(".")
                    refnames_all.append(f"{name_parts[0]}.{name_parts[-1]}")

        labels = np.array(labels_list).astype(np.float32)
        refnames_all_arr = np.array(refnames_all)

        # Build samples based on train/test split
        for idx in index:
            matches = ref_names[idx] == refnames_all_arr
            matching_indices = np.where(matches)[0].tolist()

            for img_idx in matching_indices:
                img_path = os.path.join(self.root, "dst_imgs_all", img_names[img_idx])
                for _ in range(self.patch_num):
                    self.samples.append((img_path, labels[img_idx]))


class Koniq10kFolder(BaseIQADataset):
    """
    Dataset class for the KonIQ-10k database.

    Contains 10,073 images with authentic distortions and crowd-sourced
    quality ratings.
    """

    def _build_samples(self, index: list[int]) -> None:
        """Build samples from KonIQ-10k database."""
        # Parse CSV file with image names and MOS scores
        csv_path = os.path.join(self.root, "koniq10k_scores_and_distributions.csv")
        img_names: list[str] = []
        mos_scores: list[np.floating[Any]] = []

        with open(csv_path) as f:
            reader = csv.DictReader(f)
            for row in reader:
                img_names.append(row["image_name"])
                mos = np.float32(row["MOS_zscore"])
                mos_scores.append(mos)

        # Build samples
        for idx in index:
            img_path = os.path.join(self.root, "1024x768", img_names[idx])
            for _ in range(self.patch_num):
                self.samples.append((img_path, mos_scores[idx]))


class BIDFolder(BaseIQADataset):
    """
    Dataset class for the BID (Blur Image Database).

    Contains images with realistic blur distortions.
    """

    def _build_samples(self, index: list[int]) -> None:
        """Build samples from BID database."""
        # Parse Excel file with image numbers and MOS scores
        xls_path = os.path.join(self.root, "DatabaseGrades.xlsx")
        workbook = load_workbook(xls_path)
        sheet = workbook.active

        img_names: list[str] = []
        mos_scores: list[np.floating[Any]] = []

        # Read data starting from row 2 (skip header)
        for row_idx in range(2, BID_IMAGE_COUNT + 2):
            img_num = sheet.cell(row=row_idx, column=1).value
            img_name = f"DatabaseImage{img_num:04d}.JPG"
            img_names.append(img_name)

            mos = np.float32(sheet.cell(row=row_idx, column=2).value)
            mos_scores.append(mos)

        # Build samples
        for idx in index:
            img_path = os.path.join(self.root, img_names[idx])
            for _ in range(self.patch_num):
                self.samples.append((img_path, mos_scores[idx]))


class TID2013Folder(BaseIQADataset):
    """
    Dataset class for the TID2013 database.

    Contains 25 reference images with 24 distortion types at 5 levels each.
    """

    def _build_samples(self, index: list[int]) -> None:
        """Build samples from TID2013 database."""
        # Load reference image identifiers
        ref_path = os.path.join(self.root, "reference_images")
        ref_ids = get_tid_reference_names(ref_path, ".bmp.BMP")

        # Parse label file
        txt_path = os.path.join(self.root, "mos_with_names.txt")
        img_names: list[str] = []
        labels_list: list[str] = []
        ref_ids_all: list[str] = []

        with open(txt_path, "r") as f:
            for line in f:
                parts = line.strip().split()
                if len(parts) >= 2:
                    img_names.append(parts[1])
                    labels_list.append(parts[0])
                    # Extract reference ID from image name (e.g., "i01_05_3" -> "01")
                    ref_id = parts[1].split("_")[0][1:]
                    ref_ids_all.append(ref_id)

        labels = np.array(labels_list).astype(np.float32)
        ref_ids_all_arr = np.array(ref_ids_all)

        # Build samples based on train/test split
        for idx in index:
            matches = ref_ids[idx] == ref_ids_all_arr
            matching_indices = np.where(matches)[0].tolist()

            for img_idx in matching_indices:
                img_path = os.path.join(
                    self.root, "distorted_images", img_names[img_idx]
                )
                for _ in range(self.patch_num):
                    self.samples.append((img_path, labels[img_idx]))


# Legacy function aliases for backward compatibility
def getFileName(path: str, suffix: str) -> list[str]:
    """Legacy alias for get_filenames_by_suffix."""
    return get_filenames_by_suffix(path, suffix)


def getTIDFileName(path: str, suffix: str) -> list[str]:
    """Legacy alias for get_tid_reference_names."""
    return get_tid_reference_names(path, suffix)
